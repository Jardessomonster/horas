import openpyxl
import os
from openpyxl.styles import PatternFill, Border, Side


# constants
TOTAL_HOURS_POSITION = 'I2'
TOTAL_VALUE_POSITION = 'J2'
RELATORY = 'relatorio.xlsx'
BLUE = '0000FF'
BLACK = '000000'

def get_file_owner(file):
    owner = file.split('-')[0]
    return owner

def get_file_data(file):
    wb = openpyxl.load_workbook(file, read_only = True, data_only = True)
    ws = wb[wb.sheetnames[0]]

    hours = ws[TOTAL_HOURS_POSITION].value
    value = ws[TOTAL_VALUE_POSITION].value

    wb.close()

    return { 'hours': hours, 'value': value }

def check_if_relatory_exists():
    os.chdir('../')
    path = os.getcwd()
    files = os.listdir(path)
    os.chdir('excels')

    return RELATORY in files
    
def create_relatory():
    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[0]]
    # setting values
    ws['A1'].value = 'Nome'
    ws['B1'].value = 'Horas Totais'
    ws['C1'].value = 'Valor Total'
    # setting style
    thin = Side(border_style="thin", color=BLACK)
    # border
    ws['A1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws['B1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws['C1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # bg color   
    ws['A1'].fill = PatternFill(start_color = BLUE, end_color = BLUE, fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = BLUE, end_color = BLUE, fill_type = 'solid')
    ws['C1'].fill = PatternFill(start_color = BLUE, end_color = BLUE, fill_type = 'solid')

    wb.save('../' + RELATORY)
    wb.close()

def write_data(name, data, position):
    wb = openpyxl.load_workbook('../' + RELATORY)
    ws = wb[wb.sheetnames[0]]

    ws['A{}'.format(position)].value = name
    ws['B{}'.format(position)].value = data['hours']
    ws['C{}'.format(position)].value = data['value']

    wb.save('../' + RELATORY)
    wb.close()
     
def main():
    os.chdir('excels')
    path = os.getcwd()
    files = os.listdir(path)
    # to count witch cell to write, make better later
    y_position = 2

    for file in files:
        if file.endswith('.xlsx'):
            name = get_file_owner(file)
            data = get_file_data(file)

            if not check_if_relatory_exists():
                create_relatory()
            
            write_data(name, data, y_position)
        
        y_position += 1
    
    print('ran successfully!!')

            
if __name__ == '__main__':
    main()